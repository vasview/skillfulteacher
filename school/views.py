from lib2to3.pgen2.pgen import DFAState
from multiprocessing import context
from multiprocessing.spawn import import_main_path
from django.template import RequestContext
from django.http import HttpResponse, HttpResponseRedirect, HttpResponseNotFound, Http404
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.views.generic import View, ListView, DetailView, CreateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout
from requests import request
from school.forms import *
from student.models import *
from people.models import Person
from .models import *
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Font, PatternFill


class SchoolHome(View):
    def get(self,request,*args, **kwargs):
        return render(request, 'school/index.html')

def user_login(request):
    next_page = ''
    if request.GET:
        next_page = request.GET['next']

    if request.POST:
        form = LoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['user']
            password = form.cleaned_data['password']
            
            user = authenticate(username=username, password=password)
            if user is not None and user.is_active:
                login(request, user)
                if next_page == "":
                    return redirect('home')
                else:
                    return redirect(next_page)
            else:
                form.add_error(None, 'Ошибка входа')
    else:
        form = LoginForm()
    
    return render(request, 'school/login.html', {'form': form, 'next': next_page})

def user_logout(request):
    logout(request)
    return redirect(reverse_lazy('home'))

### End points related to klasses in the School

class AllKlasses(LoginRequiredMixin, ListView):
    model = Klass
    context_object_name = 'klasses'
    template_name = 'school/all_klasses.html'
    login_url = '/login/'

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(**kwargs)
        teacher = self.request.user.teacher
        teacher_klasses = ClassroomTeacher.objects.filter(teacher=teacher)
        context['teacher_klasses'] = teacher_klasses
        return context

class ShowKlass(LoginRequiredMixin, DetailView):
    model = Klass
    content_object_name = 'klass'
    pk_url_kwarg = 'id'
    template_name = 'school/klass_details.html'
    login_url = '/login/'

    def get_klass(self):
        id = self.kwargs.get('id')
        return get_object_or_404(Klass, id=id)
    
    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(**kwargs)
        klass = self.get_klass()
        students = Student.objects.filter(klasses__id=klass.id)
        context['students'] = students
        groups = Group.objects.filter(klass_id=klass.id)
        context['klass_groups'] = groups
        # active_tab = request.REQUEST.get('active_tab','')
        # context['active_tab'] = active_tab
        return context

class AddStudentsInKlass(LoginRequiredMixin, View):
    form_class = StudentKlassForm
    template_name = 'school/add_students_klass.html'
    login_url = '/login/'

    def get_klass(self):
        id = self.kwargs.get('id')
        return get_object_or_404(Klass, id=id)

    def get_student(self, student_id):
        return get_object_or_404(Student, id = student_id)

    def get_context_data(self, **kwargs):
        klass = self.get_klass()
        context = {'klass': klass}
        students_in_klass = StudentKlass.objects.filter(klass=klass).values_list('student_id', flat=True)
        context['students'] = Student.objects.all().exclude(id__in=students_in_klass)
        return context

    def get(self, request, *args, **kwargs):
        # klass = self.get_klass()
        form = self.form_class()
        context = self.get_context_data()
        context['form'] = form
        # if context['students']:
        return render(request, self.template_name, context)
        # else:
            # return redirect('show_klass', id = klass.id)

    def post(self, request, *args, **kwargs):
        klass = self.get_klass()
        students = request.POST.getlist('students')

        if students:
            for std in students: 
                student = self.get_student(std)
                form = StudentKlassForm(request.POST, initial={'klass': klass, 'student': student})
                if form.is_valid():
                    form.save()
            return redirect('show_klass', id = klass.id)
        else:
            form = self.form_class(request.POST)
            return render(request, self.template_name, {'form': form, 'klass': klass})

class RemoveStudentFromKlass(LoginRequiredMixin, View):
    # model = StudentKlass
    # pk_url_kwarg = 'review_id'
    template_name = 'delete_confirmation.html'
    login_url = '/login/'

    def get_klass(self):
        id = self.kwargs.get('id')
        return get_object_or_404(Klass, id=id)

    def get_student(self):
        student_id = self.kwargs.get('std_id')
        return get_object_or_404(Student, id = student_id)

    def get(self, request, *args, **kwargs):
        klass = self.get_klass()
        student = self.get_student()
        if klass and student:
            context = {'title': 'Вы точно хотите удалить ученика из класса'}
            return render(request, self.template_name, context)
        else:
            return redirect('show_klass', id = klass.id)

    def post(self, request, *args, **kwargs):
        klass = self.get_klass()
        student = self.get_student()
        if klass and student:
            StudentKlass.objects.get(klass = klass, student = student).delete()

        return redirect('show_klass', id = klass.id)

# export student in a class as a list
def export_students_xls(request, id):
    klass_id = id
    klass = Klass.objects.get(pk=klass_id)
    teacher = ClassroomTeacher.objects.filter(klass_id=klass_id).last().teacher
    
    # create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Ученики"

    # write page header
    ws.cell(column=3, row=1, value="Список учащихся {} класса".format(klass.code))
    ws.cell(column=3, row=2, value='учебный год')
    ws.cell(column=2, row=3, value="Классный руководитель: {}".format(teacher.full_name))

    # insert empty row
    empty_row=[]
    for row in range(1,1):
        ws.append(empty_row)

    # prepare the main list header
    row_num = 5
    columns = ['ФИО', 'Дата рождения', 'Пол', 'Телефон', 'Домашний адрес', ]

    for col in range(1, len(columns)+1):
        ws.cell(column=col, row=row_num, value=columns[col - 1]).font = Font(bold=True)
        ws.cell(column=col, row=row_num).fill = PatternFill("solid", start_color="C4C4C4")
    
    # write the list of students into the table.
    person_ids = Student.objects.filter(klasses__id=klass_id).values_list('person_id', flat=True)
    rows = Person.objects.filter(pk__in=person_ids).values_list('full_name', 'birth_date', 'gender', 'phone', 'actual_address')
    
    for row in range(1, len(rows)+1):
        row_num += 1
        ws.append(rows[row-1])

    response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=my_class_students.xlsx'
    return response
  
# add group to klass
class AddKlassGroup(LoginRequiredMixin, View):
    form_class = AddKlassGroupForm
    login_url = '/login/'

    def get_klass(self):
        id = self.kwargs.get('id')
        return get_object_or_404(Klass, id=id)

    def post(self, request, *args, **kwargs):
        klass = self.get_klass()
        form = self.form_class(request.POST)
        if form.is_valid():
            form.save()
        
        return redirect('show_klass',  id = klass.id)

class EditKlassGroup(LoginRequiredMixin, View):
    template_name = 'school/partials/edit_klass_group_modal.html'
    login_url = '/login/'

    def get_klass(self):
        id = self.kwargs.get('id')
        return get_object_or_404(Klass, id=id)

    def get_group(self):
        group_id = self.kwargs.get('group_id')
        return get_object_or_404(Group, id = group_id)

    def get(self, request, *args, **kwargs):
        klass = self.get_klass()
        group = self.get_group()

        # ajax
        if request.headers.get('x-requested-with') == "XMLHttpRequest":
            print('ajax request')

        if klass and group:
            context = {
                'klass': klass,
                'group': group
            }
            return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        klass = self.get_klass()
        group = self.get_group()
        if klass and group:
            group.name = request.POST.get('name')
            group.save()
    
        return redirect('show_klass', id = klass.id)
    
class RemoveKlassGroup(LoginRequiredMixin, View):
    template_name = 'delete_confirmation.html'
    login_url = '/login/'

    def get_klass(self):
        id = self.kwargs.get('id')
        return get_object_or_404(Klass, id=id)

    def get_group(self):
        group_id = self.kwargs.get('group_id')
        return get_object_or_404(Group, id = group_id)

    def get(self, request, *args, **kwargs):
        klass = self.get_klass()
        group = self.get_group()
        if klass and group:
            context = {'title': 'Вы точно хотите удалить группу из класса'}
            return render(request, self.template_name, context)
        else:
            return redirect('show_klass', id = klass.id)

    def post(self, request, *args, **kwargs):
        klass = self.get_klass()
        group = self.get_group()
        if klass and group:
            group.delete()

        return redirect('show_klass', id = klass.id)

# adding studing to a group
class AddStudentsToGroup(LoginRequiredMixin, View):
    form_class = StudentGroupMemberForm
    template_name = 'school/add_group_members.html'
    login_url = '/login/'

    def get_group(self):
        id = self.kwargs.get('id')
        return get_object_or_404(Group, id=id)

    def get_student(self, student_id):
        return get_object_or_404(Student, id = student_id)

    def get_context_data(self, **kwargs):
        group = self.get_group()
        context = {'group': group}
        students_in_klass = StudentKlass.objects.filter(klass=group.klass).values_list('student_id', flat=True)
        context['students'] = Student.objects.filter(id__in=students_in_klass)
        return context

    def get(self, request, *args, **kwargs):
        # klass = self.get_klass()
        form = self.form_class()
        context = self.get_context_data()
        context['form'] = form
        # if context['students']:
        return render(request, self.template_name, context)
        # else:
            # return redirect('show_klass', id = klass.id)

    def post(self, request, *args, **kwargs):
        group = self.get_group()
        students = request.POST.getlist('students')

        if students:
            for std in students: 
                student = self.get_student(std)
                form = StudentGroupMemberForm(request.POST, initial={'group': group, 'student': student})
                if form.is_valid():
                    form.save()
            return redirect('show_klass', id = group.klass_id)
        else:
            form = self.form_class(request.POST)
            return render(request, self.template_name, {'form': form, 'group': group})

